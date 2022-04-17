from ast import alias
from openpyxl import *
from openpyxl.utils import *
from textwrap import *
from openpyxl.styles import *
from datetime import date
import random

wb = Workbook()
dest_filename = 'idf_assign.xlsx'
ws = wb.active

today = date.today()

ws.title = f'Assign Date - {today}'

fill = PatternFill(start_color='C4BD97', end_color='C4BD97', fill_type='solid') #Cell Fill Style
fc = Font(color="000000") #Font Color Style

a1 = ws['A1'] 
b1 = ws['B1'] 
c1 = ws['C1']

#Applying Styles to cells
a1.fill = fill 
b1.fill = fill
c1.fill = fill
a1.font = fc
b1.font = fc
c1.font = fc

a1 = ws['a1'] = 'Technician Assigned'
b1 = ws['b1'] = 'IDF'
c1 = ws['c1'] = 'Count'

print(ws.columns)
print(column_cells)
print(type(column_cells))

#Setting custom Column Widths
ws.column_dimensions['A'].width = 22.71
ws.column_dimensions['B'].width = 105.0 
ws.column_dimensions['C'].width = 7 

idfs = ['40-B01-A','40-B01-B', '30-B01-A', '30-B01-B', '30-B01-C', '30-L05-B', '30-L08-B', '50-B01-B1', '20-B01-A', '40-B01-BC', '40-B01-CC', '40-B01-FU', '30-B01-B', '30-L15-B', '50-B01-SPA', '140-MDF', '140-IDF', '41-B01-A', '40-B01-C', '40-B01-AC', '40-B01-DC', '50-B01-A1', '50-B01-A2', '50-B01-B2', '30-L18-B', '40-L03-A', '40-L05-A', '40-L05-B', '40-L08-A', '40-L08-B', '40-L11-A', '40-L11-B', '30-L21-B', '30-L24-B', '40-L15-B', '40-L18-A', '40-L18-B', '40-L21-A', '40-L21-B']

tech1 = []
tech2 = []
tech3 = []
tech4 = []


while len(idfs) >= 0:
    try:
        capture = random.choice(idfs)
        if len(tech1) != len(tech2):
            tech1.append(capture)
        elif len(tech2) != len(tech3):
            tech2.append(capture)
        elif len(tech3) != len(tech4):
            tech3.append(capture)
        else:
            tech4.append(capture)
        idfs.remove(capture)
    except:
        break

loT = [tech1,tech2,tech3,tech4] #List Of Techs to randomly assign

#Used to select a random set of idfs ffrom loT
t_cap = random.choice(loT)
loT.remove(t_cap)
t_cap_2 = random.choice(loT)
loT.remove(t_cap_2)
t_cap_3 = random.choice(loT)
loT.remove(t_cap_3)
t_cap_4 = random.choice(loT)
loT.remove(t_cap_4)

t1_count = len(t_cap)
t2_count = len(t_cap_2)
t3_count = len(t_cap_3)
t4_count = len(t_cap_4)

a2 = ws['a2'] = 'Corey Mckenzie'
a3 = ws['a3'] = 'Jermaine Douglas'
a4 = ws['a4'] = 'Christy Pinder'
a5 = ws['a5'] = 'Deanglo Turnquest'

#Assinging Random IDF List to Techs
b2 = ws['b2'] = f'{t_cap}'
b3 = ws['b3'] = f'{t_cap_2}'
b4 = ws['b4'] = f'{t_cap_3}'
b5 = ws['b5'] = f'{t_cap_4}'

#Assinging idf count to display how many responsible for
c2 = ws['c2'] = f'{t1_count}'
c3 = ws['c3'] = f'{t2_count}'
c4 = ws['c4'] = f'{t3_count}'
c5 = ws['c5'] = f'{t4_count}'

wb.save(filename=dest_filename)
